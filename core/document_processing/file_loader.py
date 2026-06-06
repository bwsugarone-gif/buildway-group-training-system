# -*- coding: utf-8 -*-
"""
core/document_processing/file_loader.py
Buildway AI Core — File Loading Utilities

Handles uploaded files (images, PDFs, DOCX, XLSX) and extracts content for analysis.
No Streamlit dependency — works with raw file paths.
"""

import os
import uuid
from datetime import datetime
from pathlib import Path

from PIL import Image
import pypdf

UPLOAD_DIR = Path(__file__).parent.parent.parent / "data" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".pdf", ".docx", ".xlsx"}
MAX_FILE_SIZE_MB = 50

FILE_TYPE_LABELS = {
    ".jpg": "Image",
    ".jpeg": "Image",
    ".png": "Image",
    ".pdf": "PDF",
    ".docx": "Word Document",
    ".xlsx": "Excel Spreadsheet",
}


def save_file(file_bytes: bytes, original_name: str, upload_dir: Path | None = None) -> Path:
    """Save raw file bytes to the uploads directory. Returns the saved path."""
    target_dir = upload_dir or UPLOAD_DIR
    target_dir.mkdir(parents=True, exist_ok=True)
    ext = Path(original_name).suffix.lower()
    unique_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    save_path = target_dir / unique_name
    save_path.write_bytes(file_bytes)
    return save_path


def is_allowed_file(filename: str) -> bool:
    """Check if file extension is allowed."""
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def get_file_type_label(filename: str) -> str:
    """Return a human-readable file type label."""
    ext = Path(filename).suffix.lower()
    return FILE_TYPE_LABELS.get(ext, ext.upper().lstrip("."))


def extract_pdf_text(file_path: Path) -> str:
    """Extract text content from a PDF file."""
    try:
        reader = pypdf.PdfReader(str(file_path))
        text_parts = []
        for idx, page in enumerate(reader.pages, 1):
            text = page.extract_text() or ""
            if text.strip():
                text_parts.append(f"[Page {idx}]\n{text.strip()}")
        return "\n\n".join(text_parts)
    except Exception as e:
        return f"[PDF extraction error: {e}]"


def extract_docx_text(file_path: Path) -> str:
    """Extract text from a DOCX file."""
    try:
        from docx import Document
        doc = Document(str(file_path))
        paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except Exception as e:
        return f"[DOCX extraction error: {e}]"


def extract_xlsx_text(file_path: Path) -> str:
    """Extract text from an XLSX file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    rows.append(" | ".join(cells))
            if rows:
                parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
        return "\n\n".join(parts)
    except Exception as e:
        return f"[XLSX extraction error: {e}]"


def load_file(file_path: Path) -> dict:
    """
    Load a file and extract its text content.
    Returns a dict with keys: path, type, text, error.
    """
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()
    result = {
        "path": str(file_path),
        "name": file_path.name,
        "type": get_file_type_label(file_path.name),
        "suffix": suffix,
        "text": "",
        "error": "",
        "size_bytes": file_path.stat().st_size if file_path.exists() else 0,
    }

    if not file_path.exists():
        result["error"] = "File not found"
        return result

    size_mb = result["size_bytes"] / (1024 * 1024)
    if size_mb > MAX_FILE_SIZE_MB:
        result["error"] = f"File too large ({size_mb:.1f} MB > {MAX_FILE_SIZE_MB} MB limit)"
        return result

    if suffix == ".pdf":
        result["text"] = extract_pdf_text(file_path)
    elif suffix == ".docx":
        result["text"] = extract_docx_text(file_path)
    elif suffix == ".xlsx":
        result["text"] = extract_xlsx_text(file_path)
    elif suffix in {".jpg", ".jpeg", ".png"}:
        result["text"] = ""  # Images handled by OCR engine
        result["type"] = "Image"
    else:
        result["error"] = f"Unsupported file type: {suffix}"

    return result
